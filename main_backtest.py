import pandas as pd
import yfinance as yf
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference

#아직 자동화 X, 단순히 백테스트용, 최적화를 모두 다 한다음 자동화할 예정

tickers = ["NVDA", "RBLX", "AAPL", "TSLA", "MSFT", "IBM", "005930.KS", "000660.KS", "009830.KS", "066570.KS", "005180.KS"]
# 횡보장  NVIDIA, APPLE, TESLA
# 하락장  ROBLOX, MICROSOFT, LG전자
# 상승장  IBM, 삼성, SK하이닉스, 한화솔루션, 빙그레

k = 0.7 # 최적화 시켜야 할 값. 낮으면 가짜 돌파에 취약해질 수 있지만 너무 높으면 큰 이익을 못 봄. -> 장에 따라 달리 해야 할까?
tuja = 100000 # 원금

all_summary = [] # 총 요약본 
portfolio_curve = pd.DataFrame()

writer = pd.ExcelWriter("portfolio_backtest.xlsx", engine="openpyxl")

sp500 = yf.download("^GSPC", period="365d")

if isinstance(sp500.columns, pd.MultiIndex):
    sp500.columns = sp500.columns.get_level_values(0)

sp500['MA200'] = sp500['Close'].rolling(200).mean()
sp500_filter = sp500['Close'] > sp500['MA200']


kospi = yf.download("^KS11", period="365d")

if isinstance(kospi.columns, pd.MultiIndex):
    kospi.columns = kospi.columns.get_level_values(0)

kospi['MA200'] = kospi['Close'].rolling(200).mean()
kospi_filter = kospi['Close'] > kospi['MA200']

def make_summary(df_part):
    total_trades = df_part['breakout'].sum()
    wins = df_part['Win'].sum()
    win_rate = (wins / total_trades) * 100 if total_trades > 0 else 0

    return {
        "거래 횟수": total_trades,
        "승리 횟수": wins,
        "승률(%)": win_rate,
        "총 손익": df_part['Real_Profit'].sum(),
        "누적 수익률(%)": df_part['Real_Return_%'].sum()
    }

# 종목 루프

for ticker in tickers:

    df = yf.download(ticker, period="365d")
    df.columns = df.columns.get_level_values(0)

    if ".KS" in ticker:
        market_condition = kospi_filter.reindex(df.index).fillna(False)
    else:
        market_condition = sp500_filter.reindex(df.index).fillna(False)

    market_condition = market_condition.astype(bool)

    df['range'] = df['High'].shift(1) - df['Low'].shift(1)
    df['target'] = df['Open'] + df['range'] * k
    df['MA20'] = df['Close'].rolling(20).mean()
    df['MA50'] = df['Close'].rolling(50).mean()
    df['breakout'] = (
    (df['High'] > df['target']) &
    (df['Close'] > df['MA20']) &
    (df['MA20'] > df['MA50']) &
    market_condition
)

    df['Buy_Price'] = df['target'].where(df['breakout'])
    df['Sell_Price'] = df['Close'].where(df['breakout'])

    # 투자금 반영시키기
    df['Qty'] = tuja / df['Buy_Price']
    df['Real_Profit'] = df['Qty'] * (df['Sell_Price'] - df['Buy_Price'])
    df['Real_Return_%'] = (df['Real_Profit'] / tuja) * 100

    df['Cumulative_Profit'] = df['Real_Profit'].cumsum()
    df['Cumulative_Return_%'] = df['Real_Return_%'].cumsum()

    df['Win'] = df['Real_Profit'] > 0

    # ===========================================================================================================================
    # 종목 요약
    # ===========================================================================================================================

    total_trades = df['breakout'].sum()
    wins = df['Win'].sum()
    win_rate = (wins / total_trades) * 100 if total_trades > 0 else 0

    summary_all = make_summary(df)
    summary_100 = make_summary(df.tail(100))
    summary_30 = make_summary(df.tail(30))

    summary = {
        "Ticker": ticker,

        "365일 거래": summary_all["거래 횟수"],
        "365일 승률": summary_all["승률(%)"],
        "365일 총손익": summary_all["총 손익"],

        "100일 거래": summary_100["거래 횟수"],
        "100일 승률": summary_100["승률(%)"],
        "100일 총손익": summary_100["총 손익"],

        "30일 거래": summary_30["거래 횟수"],
        "30일 승률": summary_30["승률(%)"],
        "30일 총손익": summary_30["총 손익"],
    }

    all_summary.append(summary)


    # 결과 저장

    result = df[['Open',
                 'target',
                 'breakout',
                 'Buy_Price',
                 'Sell_Price',
                 'Qty',
                 'Real_Profit',
                 'Real_Return_%',
                 'Cumulative_Profit',
                 'Cumulative_Return_%',
                 'Win']].round(2)

    result.to_excel(writer, sheet_name=ticker)

    ws = writer.sheets[ticker]

    # 열 너비 자동 조정 !! - 포트폴리오 시트에는 해당이 안됨. 수정 필요할 듯?
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # 승패 색상
    green = PatternFill(start_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            if cell.value:
                cell.fill = green
            else:
                cell.fill = red

    # 누적 수익 그래프  !! - 이거 지금 기능 좀 이상함. 수정해야 함. 나중에 보면 해라
    chart = LineChart()
    chart.title = f"{ticker} Cumulative Profit"

    data = Reference(
        ws,
        min_col=10,
        min_row=1,
        max_row=len(result) + 1
    )

    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, "M2")

    # 포트폴리오 합산용
    portfolio_curve[ticker] = df['Real_Profit'].fillna(0)

# 포트폴리오 통합

portfolio_curve['Total_Profit'] = portfolio_curve.sum(axis=1)
portfolio_curve['Cumulative'] = portfolio_curve['Total_Profit'].cumsum()

portfolio_curve.to_excel(writer, sheet_name="Portfolio")

# 요약 시트 결과는 엑셀파일에서 확인하기!!!!! -> portfolio_backtest.xlsx
summary_df = pd.DataFrame(all_summary)
summary_df.to_excel(writer, sheet_name="Summary", index=False)

writer.close()
